# AUTOMAÇÃO PARA CRIAR PLANILHA DE BOLETOS A BAIXAR E JA DAR BAIXA NOS BOLETOS RECEBIDOS
# BUSCA AS INFORMAÇÕES NO extrato banco.xlsx DOS BOLETOS QUE FORAM PAGOS E ATUALIZA O RELATORIO EM EXCEL

import re
import time
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

PASTA_SCRIPT = Path(__file__).resolve().parent
caminho_origem = PASTA_SCRIPT / 'extrato banco.xlsx'
caminho_destino = PASTA_SCRIPT / 'documentos.xlsx'
caminho_log_erro = PASTA_SCRIPT / 'erro_boletos_a_baixar.txt'
caminho_relatorio = PASTA_SCRIPT / 'Relatório de Boletos Baixados.xlsx'
caminho_relatorio_antigo = PASTA_SCRIPT / 'Relatório ee Boletos baixados.xlsx'
URL_SISTEMA = "http://192.168.0.32:82"


def mostrar_e_salvar_erro(mensagem):
    erro_completo = traceback.format_exc()
    texto_log = f"{mensagem}\n\n{erro_completo}"
    caminho_log_erro.write_text(texto_log, encoding='utf-8')
    print("\n" + "=" * 70)
    print(mensagem)
    print(f"Erro completo salvo em: {caminho_log_erro}")
    print("=" * 70)
    print(erro_completo)
    try:
        input("\nPressione ENTER para fechar...")
    except EOFError:
        pass


def switch_to_correct_frame(driver):
    driver.switch_to.default_content()
    frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")
    if not frames:
        return True if driver.find_elements(By.XPATH, "//input") else False
    for frame in frames:
        driver.switch_to.default_content()
        try:
            driver.switch_to.frame(frame)
            if driver.find_elements(By.XPATH, "//input | //button | //a"):
                return True
        except Exception:
            continue
    return False


def realizar_login(driver):
    print("Tentando acessar/logar no sistema...")
    driver.get(URL_SISTEMA)
    time.sleep(3)

    wait = WebDriverWait(driver, 20)

    try:
        driver.switch_to.default_content()
        campos_login = driver.find_elements(By.NAME, "NOME")
        if campos_login:
            print("Tela de login detectada. Inserindo credenciais...")
            campos_login[0].send_keys("roberto")
            driver.find_element(By.NAME, "SASIC").send_keys("913942")

            try:
                driver.find_element(By.XPATH, "//input[@value='Entrar']").click()
            except Exception:
                driver.find_element(By.NAME, "LOGIN").click()

            time.sleep(5)

        print("Buscando acesso ao setor 'Contas a receber'...")
        if switch_to_correct_frame(driver):
            try:
                btn_contas = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[contains(@value,'Contas a receber')]"))
                )
                btn_contas.click()
                print("Setor acessado com sucesso.")
                time.sleep(3)
                return True
            except Exception as e:
                print(f"Botao 'Contas a receber' nao ficou clicavel: {e}")
        else:
            print("Nao foi possivel localizar o frame principal do sistema.")

        return False
    except Exception as e:
        print(f"Erro durante o login/acesso: {e}")
        return False


def wait_and_click(driver, xpath, timeout=10):
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.click()
    return element


def formatar_valor_monetario(texto):
    if not texto:
        return None
    limpo = texto.replace('R$', '').replace(' ', '').strip()
    if ',' in limpo and '.' in limpo:
        limpo = limpo.replace('.', '').replace(',', '.')
    elif ',' in limpo:
        limpo = limpo.replace(',', '.')
    try:
        limpo = re.sub(r'[^0-9.]', '', limpo)
        return float(limpo)
    except Exception:
        return None


def normalizar_relatorio(df_relatorio):
    df_relatorio = df_relatorio.copy()
    df_relatorio = df_relatorio.dropna(how='all')

    if "Documento" in df_relatorio.columns:
        df_relatorio["Documento"] = (
            df_relatorio["Documento"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
        )

    if "Status" in df_relatorio.columns:
        df_relatorio["Status"] = df_relatorio["Status"].astype(str).str.strip()

    return df_relatorio


def salvar_relatorio_acumulado(resultado):
    df_novo = pd.DataFrame(resultado)
    if df_novo.empty:
        print("Nenhuma linha nova para incluir no relatorio.")
        return caminho_relatorio

    df_novo = normalizar_relatorio(df_novo)

    if caminho_relatorio_antigo.exists() and not caminho_relatorio.exists():
        caminho_relatorio_antigo.replace(caminho_relatorio)

    if caminho_relatorio.exists():
        df_existente = normalizar_relatorio(ler_relatorio_existente())
        df_relatorio = pd.concat([df_existente, df_novo], ignore_index=True)
    else:
        df_relatorio = df_novo

    linhas_antes = len(df_relatorio)
    df_relatorio = normalizar_relatorio(df_relatorio)
    df_relatorio = df_relatorio.drop_duplicates(subset=["Documento", "Status"], keep='first')
    linhas_removidas = linhas_antes - len(df_relatorio)
    escrever_relatorio_formatado(df_relatorio)

    print(f"\nPROCESSO CONCLUIDO. Planilha: {caminho_relatorio}")
    print(f"Novas linhas processadas nesta execucao: {len(df_novo)}")
    if linhas_removidas:
        print(f"Linhas duplicadas ignoradas: {linhas_removidas}")

    return caminho_relatorio


def ler_relatorio_existente():
    try:
        df_existente = pd.read_excel(caminho_relatorio, skiprows=3)
        if {"Documento", "Status"}.issubset(df_existente.columns):
            return df_existente
    except Exception:
        pass

    df_existente = pd.read_excel(caminho_relatorio)
    if {"Documento", "Status"}.issubset(df_existente.columns):
        return df_existente

    return pd.DataFrame(columns=["Documento", "Status"])


def escrever_relatorio_formatado(df_relatorio):
    with pd.ExcelWriter(caminho_relatorio, engine='openpyxl') as writer:
        df_relatorio.to_excel(writer, index=False, startrow=3, sheet_name='Boletos Baixados')

    wb = load_workbook(caminho_relatorio)
    ws = wb['Boletos Baixados']
    max_col = max(ws.max_column, 2)
    ultima_coluna = get_column_letter(max_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = "Relat\u00f3rio de Boletos Baixados"
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = GradientFill(type="linear", degree=0, stop=("17365D", "5B9BD5"))
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws["A2"] = "Hist\u00f3rico acumulado de baixas realizadas no SIC"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="1F4E79")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = GradientFill(type="linear", degree=0, stop=("EAF3F8", "D9EAD3"))
    ws.row_dimensions[2].height = 24

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    ws["A3"] = f"Total de registros: {max(ws.max_row - 4, 0)}"
    ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="44546A")
    ws["A3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["A3"].fill = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[3].height = 20

    header_fill = GradientFill(type="linear", degree=0, stop=("1F4E79", "70AD47"))
    border_color = Side(style="thin", color="D9E2F3")
    dark_border = Side(style="medium", color="1F4E79")
    header_row = 4
    for cell in ws[header_row]:
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = Border(top=dark_border, bottom=dark_border, left=border_color, right=border_color)
    ws.row_dimensions[header_row].height = 24

    status_coluna = None
    documento_coluna = None
    for cell in ws[header_row]:
        if cell.value == "Status":
            status_coluna = cell.column
        if cell.value == "Documento":
            documento_coluna = cell.column

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=max_col), start=5):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "F8FBFD")
        if status_coluna:
            status_texto = str(row[status_coluna - 1].value or "").lower()
            if "quitado" in status_texto:
                fill = PatternFill("solid", fgColor="C6EFCE")
            elif "erro" in status_texto or "divergente" in status_texto:
                fill = PatternFill("solid", fgColor="F4B183")
            elif "localizado" in status_texto:
                fill = PatternFill("solid", fgColor="FFEB9C")

        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.border = Border(top=border_color, bottom=border_color, left=border_color, right=border_color)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10, color="1F1F1F")

        if documento_coluna:
            documento_cell = row[documento_coluna - 1]
            documento_cell.alignment = Alignment(vertical="center", horizontal="center")
            documento_cell.font = Font(name="Calibri", size=10, bold=True, color="1F1F1F")

    if status_coluna:
        for coluna_status in ws.iter_cols(min_col=status_coluna, max_col=status_coluna, min_row=5, max_row=ws.max_row):
            for status_cell in coluna_status:
                status_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    for col_idx in range(1, max_col + 1):
        coluna = get_column_letter(col_idx)
        maior_texto = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[coluna].width = min(max(maior_texto + 3, 16), 48)

    if documento_coluna:
        ws.column_dimensions[get_column_letter(documento_coluna)].width = 18
    if status_coluna:
        ws.column_dimensions[get_column_letter(status_coluna)].width = 42

    ws.auto_filter.ref = f"A4:{ultima_coluna}{ws.max_row}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F4E79"
    wb.save(caminho_relatorio)

def dar_baixa_boletos(planilha_input):
    print("ROBO SICWEB INICIANDO...")

    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    options.add_argument('--ignore-certificate-errors')

    servico = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=servico, options=options)
    driver.maximize_window()

    try:
        if not realizar_login(driver):
            print("Segunda tentativa de login/estabilizacao...")
            if not realizar_login(driver):
                raise RuntimeError("Falha critica no login inicial.")

        wait = WebDriverWait(driver, 12)
        df = pd.read_excel(planilha_input)
        resultado = []

        for index, row in df.iterrows():
            documento = str(row["Documento"]).split(".")[0].strip()
            valor_excel = float(row["Valor"])
            status = "Nao processado"

            print(f"\n[{index + 1}/{len(df)}] Processando Doc: {documento}")

            try:
                if not switch_to_correct_frame(driver):
                    realizar_login(driver)
                    switch_to_correct_frame(driver)

                campos = driver.find_elements(By.XPATH, "//input[contains(@name, 'doc') or contains(@name, 'DOC')]")
                if not campos:
                    realizar_login(driver)
                    switch_to_correct_frame(driver)
                    campo_doc = wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//input[contains(@name, 'doc') or contains(@name, 'DOC')]")
                        )
                    )
                else:
                    campo_doc = campos[0]

                campo_doc.clear()
                campo_doc.send_keys(documento)
                wait_and_click(driver, "//input[contains(@value,'Pesquisar')]")

                time.sleep(2)

                texto_pagina = driver.find_element(By.TAG_NAME, "body").text
                if "nao localizado em contas a receber" in texto_pagina.lower() or "não localizado em contas a receber" in texto_pagina.lower():
                    status = (
                        "Documento ja foi quitado anteriormente"
                        if "Contas recebidas com este numero" in texto_pagina
                        or "Contas recebidas com este número" in texto_pagina
                        else "Numero de documento nao localizado"
                    )
                    try:
                        btn_v = driver.find_elements(By.XPATH, "//input[@value='Voltar']")
                        btn_v[0].click() if btn_v else driver.back()
                    except Exception:
                        driver.back()
                    resultado.append({"Documento": documento, "Status": status})
                    continue

                tabela = wait.until(EC.presence_of_element_located((By.XPATH, "//table")))
                linhas = tabela.find_elements(By.TAG_NAME, "tr")

                if len(linhas) < 2:
                    status = "Nao encontrado"
                else:
                    colunas_cabecalho = linhas[0].find_elements(By.XPATH, ".//td | .//th")
                    idx_valor, idx_acao, idx_obs = 5, 7, 8

                    for i, col_h in enumerate(colunas_cabecalho):
                        txt_h = col_h.text.upper()
                        if "VALOR" in txt_h and "PAGO" not in txt_h:
                            idx_valor = i
                        if "QUITAR" in txt_h or "ACAO" in txt_h or "AÇÃO" in txt_h:
                            idx_acao = i
                        if "OBS" in txt_h or "HIST" in txt_h or i == 8:
                            idx_obs = i

                    colunas_dados = linhas[1].find_elements(By.TAG_NAME, "td")
                    celula_valor = colunas_dados[idx_valor]
                    texto_valor_tela = celula_valor.text.strip()
                    if not texto_valor_tela:
                        inputs = celula_valor.find_elements(By.TAG_NAME, "input")
                        if inputs:
                            texto_valor_tela = inputs[0].get_attribute("value")

                    valor_tela = formatar_valor_monetario(texto_valor_tela)
                    try:
                        texto_obs = colunas_dados[idx_obs].text.strip()
                    except Exception:
                        texto_obs = ""

                    if valor_tela is None:
                        status = "Erro leitura valor"
                    elif abs(valor_tela - valor_excel) > 0.01:
                        status = f"Valor divergente (Tela: {valor_tela}). Obs: {texto_obs}"
                    else:
                        botoes = colunas_dados[idx_acao].find_elements(By.XPATH, ".//input | .//button | .//a")
                        if botoes:
                            botoes[0].click()
                            time.sleep(2)
                            confirms = driver.find_elements(
                                By.XPATH,
                                "//input[@value='Quitar'] | //input[@value='Gravar']",
                            )
                            if confirms:
                                confirms[0].click()
                                time.sleep(2.5)
                                status = "Quitado"
                                print("Sucesso")
                            else:
                                status = "Botao confirmacao ausente"
                        else:
                            status = "Botao acao ausente"

                try:
                    switch_to_correct_frame(driver)
                    btn_voltar = driver.find_elements(By.XPATH, "//input[@value='Voltar']")
                    btn_voltar[0].click() if btn_voltar else driver.back()
                except Exception:
                    pass

            except Exception as e:
                print(f"Erro no documento {documento}: {e}")
                status = "Erro tecnico"
                realizar_login(driver)

            resultado.append({"Documento": documento, "Status": status})

        salvar_relatorio_acumulado(resultado)

    finally:
        print("Navegador mantido aberto para diagnostico.")

try:
    # 1. Carregar a planilha (sem cabeçalho para garantir a contagem por índice)
    df = pd.read_excel(caminho_origem, header=None)

    # 2. Critérios de filtragem (Coluna 8 é índice 7)
    status_desejados = ['Recebido por pix', 'Marcado c/ recebido', 'Recebido por boleto']
    
    # Filtrar e garantir que não estamos trabalhando com uma "view" mas sim um novo DF
    df_filtrado = df[df.iloc[:, 7].isin(status_desejados)].copy()

    if df_filtrado.empty:
        print("Aviso: Nenhuma linha corresponde aos critérios de busca na coluna 8.")
    else:
        # 3. Selecionar apenas as colunas 3 e 7 (índices 2 e 6)
        df_final = df_filtrado.iloc[:, [2, 6]].copy()
        df_final.columns = ['Documento', 'Valor']

        # --- ETAPA DE LIMPEZA (Correção para o problema da coluna vazia) ---
        
        # Função para limpar e converter para número
        def limpar_valor(x):
            if pd.isna(x): return 0
            if isinstance(x, (int, float)):
                return float(x)
            x = str(x)
            x = x.replace('R$', '').replace(' ', '')
            x = x.replace('.', '')
            x = x.replace(',', '.')
            try:
                return float(x)
            except:
                return 0

        # Aplicar a limpeza na coluna 'valor'
        df_final['Valor'] = df_final['Valor'].apply(limpar_valor)
        
        # Garantir que 'documentos' também seja numérico (se possível)
        df_final['Documento'] = pd.to_numeric(df_final['Documento'], errors='coerce')

        # 4. Salvar o resultado
        df_final.to_excel(caminho_destino, index=False)
        print(f"Sucesso! O arquivo '{caminho_destino}' foi gerado com {len(df_final)} linhas.")
        print("Iniciando baixa dos boletos no SIC...")
        dar_baixa_boletos(caminho_destino)
        input("\nAutomacao finalizada. Pressione ENTER para fechar...")

except FileNotFoundError:
    mostrar_e_salvar_erro(f"Erro: O arquivo '{caminho_origem}' nao foi encontrado.")
    print(f"Erro: O arquivo '{caminho_origem}' não foi encontrado.")
except Exception as e:
    mostrar_e_salvar_erro(f"Ocorreu um erro: {e}")
    print(f"Ocorreu um erro: {e}")
